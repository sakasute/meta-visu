#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import json
import regex
import pprint

DATASET = 'finnish-birth-cohort'
DATA_PATH = 'public/data/' + DATASET + '/'
START_ROW = 4

REG_ADM_COL = 0  # 'A'
REG_COL = 1  # 'B'
CAT_COL = 2  # 'C'
COH_87_COL = 3  # 'D'
COH_97_COL = 4  # 'E'
NOTE_COL = 5  # 'F'

CATEGORIES = {"subjects": {"fi": "kohorttilaiset", "en": "subjects"}, "parents": {"fi": "vanhemmat", "en": "parents"}}

data = []


def parse_dates(raw_dates_str):
    dates = []
    raw_dates_str = raw_dates_str.replace(' ', '')
    dates_list = raw_dates_str.split(';')
    for date_str in dates_list:
        start_date = format_date(date_str.split('-')[0], '-01-01')
        if start_date:
            end_date = format_date(date_str.split('-')[1], '-12-31') if len(date_str.split('-')) > 1 else start_date

            dates.append({'start_date': start_date, 'end_date': end_date})
    return dates


def format_date(date_str, default_date):
    valid1 = regex.match('^[0-9]{4}$', date_str)
    valid2 = regex.match('^[0-9]{1,2}\.[0-9]{1,2}\.[0-9]{4}$', date_str)

    if valid1:
        return date_str + default_date
    elif valid2:
        date_components = date_str.split('.')
        day = date_components[0].zfill(2)
        month = date_components[1].zfill(2)
        year = date_components[2]
        return year + '-' + month + '-' + day
    else:
        return False


def find_by_name(list, name, lang):
    for idx, item in enumerate(list):
        if item['name'][lang] == name[lang]:
            return idx
    else:
        return None


def parse_sheet(sheet, sampling_category):
    register_admin = {
        'en': '',
        'fi': '',
    }
    register = {
        'en': '',
        'fi': '',
    }
    category = {
        'en': '',
        'fi': '',
    }
    iterator = sheet.iter_rows(min_row=START_ROW)
    for row in iterator:
        row_fi = row
        try:
            row_en = next(iterator)
        except StopIteration:
            break

        if len(row_en) == 0:
            break

        register_admin['fi'] = row_fi[REG_ADM_COL].value if row_fi[REG_ADM_COL].value != None else register_admin['fi']
        register_admin['en'] = row_en[REG_ADM_COL].value if row_en[REG_ADM_COL].value != None else register_admin['en']
        register_admin_idx = find_by_name(data, register_admin, 'en')

        if register_admin_idx == None:
            # using dict() here creates a copy of the dict so the function doesn't modify the original
            register_admin_idx = create_register_admin(dict(register_admin))

        register['fi'] = row_fi[REG_COL].value if row_fi[REG_COL].value != None else register['fi']
        register['en'] = row_en[REG_COL].value if row_en[REG_COL].value != None else register['en']
        register_idx = find_by_name(data[register_admin_idx]['registers'], register, 'en')

        if register_idx == None:
            # using dict() here creates a copy of the dict so the function doesn't modify the original
            register_idx = create_register(dict(register), register_admin_idx)

        category['fi'] = row_fi[CAT_COL].value if row_fi[CAT_COL].value != None else category['fi']
        category['en'] = row_en[CAT_COL].value if row_en[CAT_COL].value != None else category['en']
        category_idx = find_by_name(data[register_admin_idx]['registers'][register_idx]['categories'], category, 'en')

        if category_idx == None:
            # using dict() here creates a copy of the dict so the function doesn't modify the original
            category_idx = create_category(dict(category), register_admin_idx, register_idx)

        cohort_87_dates_str = str(row_fi[COH_87_COL].value)
        cohort_97_dates_str = str(row_fi[COH_97_COL].value)

        samplings_87 = create_samplings(cohort_87_dates_str, '1987', sampling_category)
        samplings_97 = create_samplings(cohort_97_dates_str, '1997', sampling_category)

        add_samplings(samplings_87 + samplings_97, register_admin_idx, register_idx, category_idx)


def create_register_admin(register_admin):
    data.append({
        'name': register_admin,
        'registers': []
    })

    return len(data) - 1


def create_register(register, register_admin_idx):
    # TODO: add keywords
    data[register_admin_idx]['registers'].append({
        'name': register,
        'categories': []
    })
    return len(data[register_admin_idx]['registers']) - 1


def create_category(category, register_admin_idx, register_idx):
    data[register_admin_idx]['registers'][register_idx]['categories'].append({
        'name': category,
        'samplings': []
    })
    return len(data[register_admin_idx]['registers'][register_idx]['categories']) - 1


def create_samplings(dates_str, cohort, category):
    dates = parse_dates(dates_str)
    samplings = []
    for date in dates:
        samplings.append({
            'startDate': date['start_date'],
            'endDate': date['end_date'],
            'cohort': cohort,
            'category': CATEGORIES[category]
        })
    return samplings


def add_samplings(samplings, register_admin_idx, register_idx, category_idx):
    for sampling in samplings:
        data[register_admin_idx]['registers'][register_idx]['categories'][category_idx]['samplings'].append(sampling)


workbook = load_workbook('FBC-rekisterit.xlsx', read_only=True)
subject_sheet = workbook['Kohorttilaiset']
parse_sheet(subject_sheet, 'subjects')

parent_sheet = workbook['Vanhemmat']
parse_sheet(parent_sheet, 'parents')

filenames = []
# pp = pprint.PrettyPrinter(depth=5)
# pp.pprint(data)
for dataset in data:
    filename = dataset['name']['en'] + '.json'
    with open(DATA_PATH + filename, 'w') as f:
        print('Create/update file: ' + DATA_PATH + filename)
        json.dump(dataset, f, ensure_ascii=False, default=str, indent=2)
        filenames.append(filename)

with open(DATA_PATH + 'filenames.json', 'w') as f:
    json.dump(filenames, f, ensure_ascii=False, default=str, indent=2)
